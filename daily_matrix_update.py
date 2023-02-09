#This script updates the KPI on the ROC monitors

import pandas as pd
import re
from datetime import date
from openpyxl import load_workbook


#Update DM
today = date.today()

#Loading weekly KPI from UPDF

sink = r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\bin.xlsx"
source = r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\Sensor_Uptime_Report.xlsm"

#-----------------------Loading uptime data_frame----------------------
updf = load_workbook(source, 
read_only = True,
data_only=True)
updf_sheet = updf["Metric"]

#Navigate Last KPI entry 
max_row = updf_sheet.max_row
kpi = updf_sheet.cell(row = max_row, column = 2).value

#------------------Loading Daily metric excel workbook-------------------

dm = load_workbook(sink, 
read_only = False)
dm_sheet = dm["recent"]

#copying row to set cell formatting in next entry 

dm_sheet['A9'] = dm_sheet.cell(row = 9, column = 1).value
dm_sheet['B9'] = dm_sheet.cell(row = 9, column = 2).value
dm_sheet['C9'] = dm_sheet.cell(row = 9, column = 3).value

#Updating the date column
dm_sheet['A9'] = today 

#Updating kpi from uptime data-frame to daily metric excel workbook
dm_sheet['B9'] = kpi

#Updating Daily difference 
for x in range(3, 9):
    y = x - 1
    z = x + 1
    cell_val =  f'=SUM(B{x}-B{y})'
    dm_sheet['C3'] = ''
    dm_sheet.cell(row = z, column = 3).value = cell_val

#deleting least recent entry
dm_sheet.delete_rows(2)

#save file
dm.save('newfile.xlsx')

#Execution confirmation
print("ALL_OK")