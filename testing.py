import pandas as pd
import re
from datetime import date
from openpyxl import load_workbook


# xl = win32com.client.Dispatch("Excel.Application") #instantiate excel app
# wb = xl.Workbooks.Open(r'C:\Users\Lesley Chingwena\Documents\Python Scripts\docs\Sensor_Uptime_Report_19_Jan_2023.xlsm')
# xl.Application.Run("Sensor_Uptime_Report_17_Jan_2023.xlsm!CommandButton1_Click()")
# wb.Save()
# xl.Application.Quit()
# print('ALL_OK')

#Update DM
today = date.today()
df_today = pd.DataFrame({'Data': [today]})
str_today = str(today.strftime("%d-%b-%Y"))
date_elem = [int(date_elem) for date_elem in re.findall('[0-9]+', str_today)]

month = str_today[3] + str_today[4] + str_today[5]
day = date_elem[0]
year = date_elem[1]

#Loading weekly KPI from UPDF
#source = r"C:\Users\Lesley Chingwena\Documents\Python Scripts\docs\source.xlsm"
sink = r"C:\Users\Lesley Chingwena\Documents\Python Scripts\docs\bin.xlsx"

url = r"https://s3grp.sharepoint.com/:x:/g/EZZIMlS0z4xGi5WW-cMZ8rAByDvgRlozNydFM52Ix5Z_bw?e=tGk1jl"
source = wget.download(url)

#-----------------------Loading uptime data-frame----------------------
updf = load_workbook(source, 
read_only = True,
data_only=True)
updf_sheet = updf["Metric"]

#Navigate Last kpi entry 
wb = updf.active
max_row = updf_sheet.max_row
kpi = updf_sheet.cell(row = max_row, column = 2).value

#------------------Loading Daily metric excel workbook-------------------

dm = load_workbook(sink, 
read_only = False)
dm_sheet = dm["recent"]


#Updating the date column
dm_sheet.cell(row = 9, column = 1).value = today 

#Updating kpi from uptime data-frame to daily metric excel workbook
dm_sheet.cell(row = 9, column = 2).value = kpi

#Updating Daily difference 
for x in range(3, 9):
    y = x - 1
    z = x + 1
    cell_val =  f'=SUM(B{x}-B{y})'
    dm_sheet.cell(row = 3, column = 3).value = ""
    dm_sheet.cell(row = z, column = 3).value = cell_val


#deleting least recent entry
dm_sheet.delete_rows(2)


dm.save('newfile.xlsx')

print("ALL_OK")

# data = df.drop(columns = df.columns[1])
# df.to_csv(r"C:\Users\Lesley Chingwena\Documents\Uptime\bin.csv")
