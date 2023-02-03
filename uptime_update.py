import openpyxl
import csv
import win32com.client
#from pywinauto.application import Application

sink = r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\Sensor_Uptime_Report_31_Jan 2023.xlsm"
source = r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\uptime_2023_02_02.csv"

# Create the source workbook
source_wb = openpyxl.Workbook()

# Select the source worksheet
source_sheet = source_wb.active

# Open CSV and read it
with open(source) as file:
    reader = csv.reader(file)
    data = [row for row in reader]

# Add rows to worksheet
for row in data:
    source_sheet.append(row)

# Load the destination workbook
destination_wb = openpyxl.load_workbook(sink)
destination_sheet = destination_wb.active


# Copy data from A2 to Q735 in the source workbook to the destination workbook
for row in range(2, 736):
    for col in range(1, 17):
        destination_sheet.cell(row=row-1, column=col).value = source_sheet.cell(row=row, column=col).value

destination_wb.save(r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\Sensor_Uptime_Report_02_Feb 2023.xlsx")

# Load the Excel application
excel = win32com.client.Dispatch("Excel.Application")

wb = excel.Workbooks.Open(r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\Sensor_Uptime_Report_02_Feb 2023.xlsx")

# Save the workbook as .xlsm
wb.SaveAs(r"C:\Users\Lesley Chingwena\Documents\python_scripts\Uptime\docs\Sensor_Uptime_Report_03_Feb 2023.xlsm", FileFormat=52)

# Close the workbook
wb.Close()

# Quit the Excel application
excel.Quit()

print("ALL_OK")

# VBA application
# app = Application().start("path/to/your/VBA/application.xlsm")

# # Wait for the window to open
# app.window(title="Copy New Uptime Data").wait("visible", timeout=60)

# # Find the button and click it
# app.window(title="Copy New Uptime Data").Button0.click()

# # Save the changes to the destination workbook
# destination_wb.save("Sensor_Uptime_Report_02_Feb_2023.xlsm")